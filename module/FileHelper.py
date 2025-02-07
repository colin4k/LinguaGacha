import os
import re
import random
from datetime import datetime

import openpyxl
import rapidjson as json
from openpyxl import Workbook
from openpyxl.utils import escape

from base.Base import Base
from module.Cache.CacheItem import CacheItem
from module.Cache.CacheProject import CacheProject
from module.TextHelper import TextHelper

class FileHelper(Base):

    # 读
    def read_from_path(path: str) -> tuple[CacheProject, list[CacheItem]]:
        project = CacheProject({
            "id": f"{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(100000, 999999)}",
        })

        items = []
        items.extend(FileHelper.read_from_path_txt(path))
        items.extend(FileHelper.read_from_path_tpp(path))
        items.extend(FileHelper.read_from_path_kvjson(path))
        items.extend(FileHelper.read_from_path_messagejson(path))

        return project, items

    # 写
    def write_to_path(path: str, items: list[CacheItem]) -> None:
        FileHelper.write_to_path_txt(path, items)
        FileHelper.write_to_path_tpp(path, items)
        FileHelper.write_to_path_kvjson(path, items)
        FileHelper.write_to_path_messagejson(path, items)

    # TXT
    def read_from_path_txt(path: str) -> list[CacheItem]:
        items = []
        for root, _, files in os.walk(path):
            for file in [file for file in files if file.endswith(".txt")]:
                row = 0
                file_path = os.path.join(root, file)
                with open(file_path, "r", encoding = "utf-8") as reader:
                    for src in reader.readlines():
                        # 跳过读取失败的行
                        if src is None:
                            continue
                        else:
                            src = str(src).removesuffix("\n")

                        items.append(
                            CacheItem({
                                "src": src,
                                "dst": src,
                                "row": row,
                                "file_type": CacheItem.FileType.TXT,
                                "file_path": os.path.relpath(file_path, path),
                            })
                        )
                        row = row + 1

        return items

    # TXT
    def write_to_path_txt(path: str, items: list[CacheItem]) -> None:
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.TXT
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for file_path, items in data.items():
            output_path = os.path.join(path, file_path)
            os.makedirs(os.path.dirname(output_path), exist_ok = True)
            with open(output_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join([item.get_dst() for item in items]))

    # T++
    def read_from_path_tpp(path: str) -> list[CacheItem]:
        items = []
        for root, _, files in os.walk(path):
            target = [
                file for file in files
                if file.endswith(".xlsx")
            ]

            for file in target:
                file_path = os.path.join(root, file)
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # 跳过空表格
                if sheet.max_column == 0:
                    continue

                # 跳过不包含 T++ 表头的表格
                if sheet.cell(row = 1, column = 1).value != "Original Text":
                    continue

                for row in range(2, sheet.max_row + 1):
                    src = sheet.cell(row = row, column = 1).value
                    dst = sheet.cell(row = row, column = 2).value

                    # 跳过读取失败的行
                    if src is None:
                        continue
                    else:
                        src = str(src)

                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst if dst != None else src,
                            "row": row,
                            "file_type": CacheItem.FileType.TPP,
                            "file_path": os.path.relpath(file_path, path),
                        })
                    )

        return items

    # T++
    def write_to_path_tpp(path: str, items: list[CacheItem]) -> None:
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.TPP
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for file_path, items in data.items():
            # 新建工作表
            work_book = Workbook()
            active_sheet = work_book.active
            active_sheet.append(
                (
                    "Original Text",
                    "Initial",
                    "Machine translation",
                    "Better translation",
                    "Best translation",
                )
            )

            # 将数据写入工作表
            for item in items:
                src: str = re.sub(r"^=", " =", item.get_src())
                dst: str = re.sub(r"^=", " =", item.get_dst())
                row: int = item.get_row()

                # 如果文本是以 = 开始，则加一个空格
                # 因为 = 开头会被识别成 Excel 公式导致 T++ 导入时 卡住
                # 加入空格后，虽然还是不能直接导入 T++ ，但是可以手动复制粘贴
                try:
                    active_sheet.cell(row = row, column = 1).value = src
                except:
                    active_sheet.cell(row = row, column = 1).value = escape(src)
                try:
                    active_sheet.cell(row = row, column = 2).value = dst
                except:
                    active_sheet.cell(row = row, column = 2).value = escape(dst)

            # 保存工作簿
            output_path = os.path.join(path, file_path)
            os.makedirs(os.path.dirname(output_path), exist_ok = True)
            work_book.save(output_path)

    # kv json
    def read_from_path_kvjson(path: str) -> list[CacheItem]:
        # {
        #     "「あ・・」": "「あ・・」",
        #     "「ごめん、ここ使う？」": "「ごめん、ここ使う？」",
        #     "「じゃあ・・私は帰るね」": "「じゃあ・・私は帰るね」",
        # }

        items = []
        for root, _, files in os.walk(path):
            for file in [file for file in files if file.endswith(".json")]:
                row = 0
                file_path = os.path.join(root, file)
                with open(file_path, "r", encoding = "utf-8") as reader:
                    json_data: dict[str, str] = TextHelper.safe_load_json_dict(reader.read().strip())

                    # 格式校验
                    if json_data == {}:
                        continue

                    # 读取数据
                    for k, v in json_data.items():
                        # 格式校验
                        if not isinstance(k, str) or not isinstance(v, str):
                            continue

                        if k != "":
                            items.append(
                                CacheItem({
                                    "src": k,
                                    "dst": v,
                                    "row": row,
                                    "file_type": CacheItem.FileType.KVJSON,
                                    "file_path": os.path.relpath(file_path, path),
                                })
                            )
                            row = row + 1

        return items

    # kv json
    def write_to_path_kvjson(path: str, items: list[CacheItem]) -> None:
        # {
        #     "「あ・・」": "「あ・・」",
        #     "「ごめん、ここ使う？」": "「ごめん、ここ使う？」",
        #     "「じゃあ・・私は帰るね」": "「じゃあ・・私は帰るね」",
        # }

        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.KVJSON
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for file_path, items in data.items():
            output_path = os.path.join(path, file_path)
            os.makedirs(os.path.dirname(output_path), exist_ok = True)
            with open(output_path, "w", encoding = "utf-8") as writer:
                writer.write(
                    json.dumps(
                        {
                            item.get_src(): item.get_dst() for item in items
                        },
                        indent = 4,
                        ensure_ascii = False,
                    )
                )

    # message json
    def read_from_path_messagejson(path: str) -> list[CacheItem]:
        # [
        #     {
        #         "message": "<fgName:pipo-fog004><fgLoopX:1><fgLoopY:1><fgSx:-2><fgSy:0.5>"
        #     },
        #     {
        #         "message": "エンディングを変更しますか？"
        #     },
        #     {
        #         "message": "はい"
        #     },
        # ]

        items = []
        for root, _, files in os.walk(path):
            for file in [file for file in files if file.endswith(".json")]:
                row = 0
                file_path = os.path.join(root, file)
                with open(file_path, "r", encoding = "utf-8") as reader:
                    json_data: list[dict] = TextHelper.safe_load_json_list(reader.read().strip())

                    # 格式校验
                    if json_data == [] or not isinstance(json_data[0], dict):
                        continue

                    for v in json_data:
                        # 格式校验
                        if "message" not in v:
                            continue

                        if v.get("message") != "":
                            items.append(
                                CacheItem({
                                    "src": v.get("message"),
                                    "dst": v.get("message"),
                                    "extra_field_src": v.get("name", ""),
                                    "extra_field_dst": v.get("name", ""),
                                    "row": row,
                                    "file_type": CacheItem.FileType.MESSAGEJSON,
                                    "file_path": os.path.relpath(file_path, path),
                                })
                            )
                            row = row + 1

        return items

    # message json
    def write_to_path_messagejson(path: str, items: list[CacheItem]) -> None:
        # [
        #     {
        #         "message": "<fgName:pipo-fog004><fgLoopX:1><fgLoopY:1><fgSx:-2><fgSy:0.5>"
        #     },
        #     {
        #         "message": "エンディングを変更しますか？"
        #     },
        #     {
        #         "message": "はい"
        #     },
        # ]

        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.MESSAGEJSON
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for file_path, items in data.items():
            output_path = os.path.join(path, file_path)
            os.makedirs(os.path.dirname(output_path), exist_ok = True)

            result = []
            for item in items:
                result.append({
                    "name": item.get_extra_field_dst(),
                    "message": item.get_dst(),
                })

            with open(output_path, "w", encoding = "utf-8") as writer:
                writer.write(json.dumps(result, indent = 4, ensure_ascii = False))