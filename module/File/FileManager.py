import os
import re
import copy
import random
import shutil
import zipfile
from datetime import datetime

import openpyxl
import rapidjson as json
from bs4 import BeautifulSoup
from lxml import etree
from openpyxl import Workbook
from openpyxl.utils import escape

from base.Base import Base
from module.Cache.CacheItem import CacheItem
from module.Cache.CacheProject import CacheProject
from module.Localizer.Localizer import Localizer

class FileManager(Base):

    # 显式引用以避免打包问题
    etree

    # EPUB 文件中读取的标签范围
    EPUB_TAGS = ("p", "h1", "h2", "h3", "h4", "h5", "h6", "li","th","td","div")
    
    # 匹配 RenPy 文本的规则
    RE_RENPY = re.compile(r"\"(.*?)(?<!\\)\"(?!\")", flags = re.IGNORECASE)

    # 添加用于匹配 Markdown 图片语法的正则表达式
    RE_MD_IMAGE = re.compile(r"!\[.*?\]\(.*?\)")

    # 修正用于匹配数学公式的正则表达式 - 先匹配双美元符号再匹配单美元符号
    RE_MATH_FORMULA = re.compile(r"\$\$(.*?)\$\$|\$(.*?)\$", re.DOTALL)

    def __init__(self, config: dict) -> None:
        super().__init__()

        # 初始化
        self.input_path: str = config.get("input_folder")
        self.output_path: str = config.get("output_folder")
        self.source_language: str = config.get("source_language")
        self.target_language: str = config.get("target_language")

    # 读
    def read_from_path(self) -> tuple[CacheProject, list[CacheItem]]:
        project: CacheProject = CacheProject({
            "id": f"{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(100000, 999999)}",
        })

        items: list[CacheItem] = []
        try:
            if os.path.isfile(self.input_path):
                paths = [self.input_path]
            elif os.path.isdir(self.input_path):
                paths = [os.path.join(root, file).replace("\\", "/") for root, _, files in os.walk(self.input_path) for file in files]
            else:
                paths: list[str] = []

            items.extend(self.read_from_path_md(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".md")]))
            items.extend(self.read_from_path_txt(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".txt")]))
            items.extend(self.read_from_path_ass(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".ass")]))
            items.extend(self.read_from_path_srt(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".srt")]))
            items.extend(self.read_from_path_xlsx(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".xlsx")]))
            items.extend(self.read_from_path_epub(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".epub")]))
            items.extend(self.read_from_path_renpy(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".rpy")]))
            items.extend(self.read_from_path_kvjson(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".json")]))
            items.extend(self.read_from_path_messagejson(self.input_path, self.output_path, [path for path in paths if path.lower().endswith(".json")]))
        except Exception as e:
            self.error(f"{Localizer.get().log_read_file_fail}", e)

        return project, items

    # 写
    def write_to_path(self, items: list[CacheItem]) -> None:
        try:
            self.write_to_path_md(self.input_path, self.output_path, items)
            self.write_to_path_txt(self.input_path, self.output_path, items)
            self.write_to_path_ass(self.input_path, self.output_path, items)
            self.write_to_path_srt(self.input_path, self.output_path, items)
            self.write_to_path_xlsx(self.input_path, self.output_path, items)
            self.write_to_path_epub(self.input_path, self.output_path, items)
            self.write_to_path_renpy(self.input_path, self.output_path, items)
            self.write_to_path_kvjson(self.input_path, self.output_path, items)
            self.write_to_path_messagejson(self.input_path, self.output_path, items)
        except Exception as e:
            self.error(f"{Localizer.get().log_write_file_fail}", e)

    # 在扩展名前插入文本
    def insert_target(self, path: str) -> str:
        root, ext = os.path.splitext(path)
        return f"{root}.{self.target_language.lower()}{ext}"

    # 在扩展名前插入文本
    def insert_source_target(self, path: str) -> str:
        root, ext = os.path.splitext(path)
        return f"{root}.{self.source_language.lower()}.{self.target_language.lower()}{ext}"

    # MD
    def read_from_path_md(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                lines = [line.removesuffix("\n") for line in reader.readlines()]
                
                # 跟踪代码块状态
                in_code_block = False
                
                for line in lines:
                    # 检查是否是代码块的开始或结束
                    if line.startswith("```"):
                        in_code_block = not in_code_block
                        
                    # 确定是否应该排除此行（不翻译）
                    should_exclude = (
                        in_code_block or                             # 在代码块内
                        line.startswith("```") or                    # 代码块标记行
                        line.startswith("!(data:image/jpeg;base64") or  # base64图片数据
                        self.RE_MD_IMAGE.search(line) is not None or    # 图片语法 ![](...)
                        self.RE_MATH_FORMULA.search(line) is not None   # 数学公式 $$ ... $$ 或 $ ... $
                    )
                    
                    status = Base.TranslationStatus.EXCLUDED if should_exclude else Base.TranslationStatus.UNTRANSLATED
                    
                    items.append(
                        CacheItem({
                            "src": line,
                            "dst": line,
                            "row": len(items),
                            "file_type": CacheItem.FileType.MD,
                            "file_path": rel_path,
                            "text_type": CacheItem.TextType.MD,
                            "status": status,  # 设置翻译状态
                        })
                    )

        return items

    # MD
    def write_to_path_md(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.MD
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(self.insert_target(abs_path), "w", encoding = "utf-8") as writer:
                writer.write("\n".join([item.get_dst() for item in items]))

    # TXT
    def read_from_path_txt(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                for line in [line.removesuffix("\n") for line in reader.readlines()]:
                    items.append(
                        CacheItem({
                            "src": line,
                            "dst": line,
                            "row": len(items),
                            "file_type": CacheItem.FileType.TXT,
                            "file_path": rel_path,
                        })
                    )

        return items

    # TXT
    def write_to_path_txt(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.TXT
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(self.insert_target(abs_path), "w", encoding = "utf-8") as writer:
                writer.write("\n".join([item.get_dst() for item in items]))

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            abs_path = f"{output_path}/{Localizer.get().path_bilingual}/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(self.insert_source_target(abs_path), "w", encoding = "utf-8") as writer:
                writer.write("\n".join([
                    item.get_src() if item.get_dst() == item.get_src() else f"{item.get_src()}\n{item.get_dst()}"
                    for item in items
                ]))

    # ASS
    def read_from_path_ass(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        # [Script Info]
        # ; This is an Advanced Sub Station Alpha v4+ script.
        # Title:
        # ScriptType: v4.00+
        # PlayDepth: 0
        # ScaledBorderAndShadow: Yes

        # [V4+ Styles]
        # Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
        # Style: Default,Arial,20,&H00FFFFFF,&H0000FFFF,&H00000000,&H00000000,0,0,0,0,100,100,0,0,1,1,1,2,10,10,10,1

        # [Events]
        # Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
        # Dialogue: 0,0:00:08.12,0:00:10.46,Default,,0,0,0,,にゃにゃにゃ
        # Dialogue: 0,0:00:14.00,0:00:15.88,Default,,0,0,0,,えーこの部屋一人で使\Nえるとか最高じゃん
        # Dialogue: 0,0:00:15.88,0:00:17.30,Default,,0,0,0,,えるとか最高じゃん

        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                lines = [line.strip() for line in reader.readlines()]

                # 格式字段的数量
                in_event = False
                format_field_num = -1
                for line in lines:
                    # 判断是否进入事件块
                    if line == "[Events]":
                        in_event = True
                    # 在事件块中寻找格式字段
                    if in_event == True and line.startswith("Format:"):
                        format_field_num = len(line.split(",")) - 1
                        break

                for line in lines:
                    content = ",".join(line.split(",")[format_field_num:]) if line.startswith("Dialogue:") else ""
                    extra_field = line.replace(f"{content}", "{{CONTENT}}") if content != "" else line

                    # 添加数据
                    items.append(
                        CacheItem({
                            "src": content.replace("\\N", "\n"),
                            "dst": content.replace("\\N", "\n"),
                            "extra_field": extra_field,
                            "row": len(items),
                            "file_type": CacheItem.FileType.ASS,
                            "file_path": rel_path,
                        })
                    )

        return items

    # ASS
    def write_to_path_ass(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # [Script Info]
        # ; This is an Advanced Sub Station Alpha v4+ script.
        # Title:
        # ScriptType: v4.00+
        # PlayDepth: 0
        # ScaledBorderAndShadow: Yes

        # [V4+ Styles]
        # Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
        # Style: Default,Arial,20,&H00FFFFFF,&H0000FFFF,&H00000000,&H00000000,0,0,0,0,100,100,0,0,1,1,1,2,10,10,10,1

        # [Events]
        # Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
        # Dialogue: 0,0:00:08.12,0:00:10.46,Default,,0,0,0,,にゃにゃにゃ
        # Dialogue: 0,0:00:14.00,0:00:15.88,Default,,0,0,0,,えーこの部屋一人で使\Nえるとか最高じゃん
        # Dialogue: 0,0:00:15.88,0:00:17.30,Default,,0,0,0,,えるとか最高じゃん

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.ASS
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append(item.get_extra_field().replace("{{CONTENT}}", item.get_dst().replace("\n", "\\N")))

            with open(self.insert_target(abs_path), "w", encoding = "utf-8") as writer:
                writer.write("\n".join(result))

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            result = []
            for item in items:
                result.append(
                    item.get_extra_field().replace("{{CONTENT}}", "{{CONTENT}}\\N{{CONTENT}}")
                                          .replace("{{CONTENT}}", item.get_src().replace("\n", "\\N"), 1)
                                          .replace("{{CONTENT}}", item.get_dst().replace("\n", "\\N"), 1)
                )

            abs_path = f"{output_path}/双语对照/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(self.insert_source_target(abs_path), "w", encoding = "utf-8") as writer:
                writer.write("\n".join(result))

    # SRT
    def read_from_path_srt(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        # 1
        # 00:00:08,120 --> 00:00:10,460
        # にゃにゃにゃ

        # 2
        # 00:00:14,000 --> 00:00:15,880
        # えーこの部屋一人で使

        # 3
        # 00:00:15,880 --> 00:00:17,300
        # えるとか最高じゃん

        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                chunks = re.split(r"\n{2,}", reader.read().strip())
                for chunk in chunks:
                    lines = [line.strip() for line in chunk.splitlines()]

                    # isdecimal
                    # 字符串中的字符是否全是十进制数字。也就是说，只有那些在数字系统中被认为是"基本"的数字字符（0-9）才会返回 True。
                    # isdigit
                    # 字符串中的字符是否都是数字字符。它不仅检查十进制数字，还包括其他可以表示数字的字符，如数字上标、罗马数字、圆圈数字等。
                    # isnumeric
                    # 字符串中的字符是否表示任何类型的数字，包括整数、分数、数字字符的变种（比如上标、下标）以及其他可以被认为是数字的字符（如中文数字）。

                    # 格式校验
                    if len(lines) < 3 or not lines[0].isdecimal():
                        continue

                    # 添加数据
                    if lines[-1] != "":
                        items.append(
                            CacheItem({
                                "src": "\n".join(lines[2:]),            # 如有多行文本则用换行符拼接
                                "dst": "\n".join(lines[2:]),            # 如有多行文本则用换行符拼接
                                "extra_field": lines[1],
                                "row": str(lines[0]),
                                "file_type": CacheItem.FileType.SRT,
                                "file_path": rel_path,
                            })
                        )

        return items

    # SRT
    def write_to_path_srt(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # 1
        # 00:00:08,120 --> 00:00:10,460
        # にゃにゃにゃ

        # 2
        # 00:00:14,000 --> 00:00:15,880
        # えーこの部屋一人で使

        # 3
        # 00:00:15,880 --> 00:00:17,300
        # えるとか最高じゃん

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.SRT
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append([
                    item.get_row(),
                    item.get_extra_field(),
                    item.get_dst(),
                ])

            with open(self.insert_target(abs_path), "w", encoding = "utf-8") as writer:
                for item in result:
                    writer.write("\n".join(item))
                    writer.write("\n\n")

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            result = []
            for item in items:
                result.append([
                    item.get_row(),
                    item.get_extra_field(),
                    f"{item.get_src()}\n{item.get_dst()}",
                ])

            abs_path = f"{output_path}/双语对照/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(self.insert_source_target(abs_path), "w", encoding = "utf-8") as writer:
                for item in result:
                    writer.write("\n".join(item))
                    writer.write("\n\n")

    # XLSX
    def read_from_path_xlsx(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            wb = openpyxl.load_workbook(abs_path)
            sheet = wb.active

            # 跳过空表格
            if sheet.max_column == 0:
                continue

            for row in range(1, sheet.max_row + 1):
                src = sheet.cell(row = row, column = 1).value
                dst = sheet.cell(row = row, column = 2).value

                # 跳过读取失败的行
                if src is None:
                    continue

                # 根据是否已存在翻译数据来添加
                if dst is None:
                    items.append(
                        CacheItem({
                            "src": str(src),
                            "dst": str(src),
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.UNTRANSLATED,
                        })
                    )
                else:
                    items.append(
                        CacheItem({
                            "src": str(src),
                            "dst": str(dst),
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.EXCLUDED,
                        })
                    )

        return items

    # XLSX
    def write_to_path_xlsx(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.XLSX
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            # 新建工作表
            work_book = Workbook()
            active_sheet = work_book.active

            # 将数据写入工作表
            for item in items:
                src: str = re.sub(r"^=", " =", item.get_src())
                dst: str = re.sub(r"^=", " =", item.get_dst())
                row: int = item.get_row()

                # 如果文本是以 = 开始，则加一个空格
                # 因为 = 开头会被识别成 Excel 公式导致 T++ 导入时 卡住
                # 加入空格后，虽然还是不能直接导入 T++ ，但是可以手动复制粘贴
                try:
                    active_sheet.cell(row = row, column = 1).value = src
                except:
                    active_sheet.cell(row = row, column = 1).value = escape(src)
                try:
                    active_sheet.cell(row = row, column = 2).value = dst
                except:
                    active_sheet.cell(row = row, column = 2).value = escape(dst)

            # 保存工作簿
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            work_book.save(abs_path)

    # EPUB
    def read_from_path_epub(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 将原始文件复制一份
            os.makedirs(os.path.dirname(f"{output_path}/cache/temp/{rel_path}"), exist_ok = True)
            shutil.copy(abs_path, f"{output_path}/cache/temp/{rel_path}")

            # 数据处理
            

            with zipfile.ZipFile(abs_path, "r") as zip_reader:
                
                for path in [p for p in zip_reader.namelist() if p.lower().endswith((".html", ".xhtml"))]:
                    if path.lower().endswith((".html", ".xhtml")):
                        with zip_reader.open(path) as reader:
                            bs = BeautifulSoup(reader.read().decode("utf-8-sig"), "html.parser")

                            for dom in bs.find_all(FileManager.EPUB_TAGS):
                                # 跳过空标签或嵌套标签
                                if dom.get_text().strip() == "" or dom.find(FileManager.EPUB_TAGS) != None:
                                    continue

                                # 提取文本 - 保留完整HTML结构

                                # 非导航页：保存整个DOM元素内容
                                items.append(
                                    CacheItem({
                                        "src": str(dom),  # 保存整个带标签的元素内容，包括子标签
                                        "dst": str(dom),  # 起始时译文与原文相同
                                        "tag": path,     # 记录标签路径以便后续处理
                                        "row": len(items),
                                        "file_type": CacheItem.FileType.EPUB,
                                        "file_path": rel_path,
                                    })
                                )
                                
                    elif path.lower().endswith(".ncx"):
                        with zip_reader.open(path) as reader:
                            bs = BeautifulSoup(reader.read().decode("utf-8-sig"), "lxml-xml")
                            for dom in bs.find_all("text"):
                                # 跳过空标签
                                if dom.get_text() == "":
                                    continue

                                items.append(CacheItem({
                                    "src": dom.get_text(),
                                    "dst": dom.get_text(),
                                    "tag": path,
                                    "row": len(items),
                                    "file_type": CacheItem.FileType.EPUB,
                                    "file_path": rel_path,
                                }))

        return items

    # EPUB
    def write_to_path_epub(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:

        def process_opf(zip_reader: zipfile.ZipFile, path: str) -> None:
            with zip_reader.open(path) as reader:
                zip_writer.writestr(
                    path,
                    reader.read().decode("utf-8-sig").replace("page-progression-direction=\"rtl\"", ""),
                )

        def process_css(zip_reader: zipfile.ZipFile, path: str) -> None:
            with zip_reader.open(path) as reader:
                zip_writer.writestr(
                    path,
                    re.sub(r"[^;\s]*writing-mode\s*:\s*vertical-rl;*", "", reader.read().decode("utf-8-sig")),
                )

        def process_ncx(zip_reader: zipfile.ZipFile, path: str, items: list[CacheItem]) -> None:
            with zip_reader.open(path) as reader:
                target = [item for item in items if item.get_tag() == path]
                bs = BeautifulSoup(reader.read().decode("utf-8-sig"), "lxml-xml")
                for dom in bs.find_all("text"):
                    # 跳过空标签
                    if dom.get_text().strip() == "":
                        continue

                    # 处理不同情况
                    item = target.pop(0)
                    dom_a = dom.find("a")
                    if dom_a != None:
                        dom_a.string = item.get_dst()
                    else:
                        dom.string = item.get_dst()

                # 将修改后的内容写回去
                zip_writer.writestr(path, str(bs))

        def process_html(zip_reader: zipfile.ZipFile, path: str, items: list[CacheItem], bilingual: bool) -> None:
            with zip_reader.open(path) as reader:
                target = [item for item in items if item.get_tag() == path]
                bs = BeautifulSoup(reader.read().decode("utf-8-sig"), "html.parser")

                # 判断是否是导航页
                is_nav_page = bs.find("nav", attrs = {"epub:type": "toc"}) != None

                # 移除竖排样式
                for dom in bs.find_all():
                    class_content: str = re.sub(r"[hv]rtl|[hv]ltr", "", " ".join(dom.get("class", "")))
                    if class_content == "":
                        dom.attrs.pop("class", None)
                    else:
                        dom["class"] = class_content.split(" ")
                    style_content: str = re.sub(r"[^;\s]*writing-mode\s*:\s*vertical-rl;*", "", dom.get("style", ""))
                    if style_content == "":
                        dom.attrs.pop("style", None)
                    else:
                        dom["style"] = style_content

                for dom in bs.find_all(FileManager.EPUB_TAGS):
                    # 跳过空标签或嵌套标签
                    if dom.get_text().strip() == "" or dom.find(FileManager.EPUB_TAGS) != None:
                        continue

                    # 取数据
                    item = target.pop(0)

                    # 输出双语，但是避免重复的行
                    if bilingual == True and item.get_src() != item.get_dst():
                        line_src = copy.copy(dom)
                        line_src["style"] = line_src.get("style", "").removesuffix(";") + "opacity:0.50;"
                        dom.insert_before(line_src)
                        dom.insert_before("\n")

                    # 根据不同类型的页面处理不同情况
                    if is_nav_page == False:
                        # 直接用翻译后的HTML内容替换整个DOM
                        # 这里item.get_dst()已经包含了完整的HTML标签结构
                        new_dom = BeautifulSoup(item.get_dst(), "html.parser")
                        # 确保解析成功且有内容
                        if new_dom and len(new_dom.contents) > 0:
                            dom.replace_with(new_dom)
                    elif item.get_src() in dom.get_text():
                        dom.replace_with(BeautifulSoup(str(dom).replace(item.get_src(), item.get_dst()), "html.parser"))
                    else:
                        pass

                # 将修改后的内容写回去
                zip_writer.writestr(path, str(bs))

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.EPUB
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            # 按行号排序
            items = sorted(items, key = lambda x: x.get_row())

            # 数据处理
            abs_path = f"{output_path}/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with zipfile.ZipFile(self.insert_target(abs_path), "w") as zip_writer:
                with zipfile.ZipFile(f"{output_path}/cache/temp/{rel_path}", "r") as zip_reader:
                    for path in zip_reader.namelist():
                        if path.lower().endswith(".css"):
                            process_css(zip_reader, path)
                        elif path.lower().endswith(".opf"):
                            process_opf(zip_reader, path)
                        elif path.lower().endswith(".ncx"):
                            process_ncx(zip_reader, path, items)
                        elif path.lower().endswith((".html", ".xhtml")):
                            process_html(zip_reader, path, items, False)
                        else:
                            zip_writer.writestr(path, zip_reader.read(path))

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            # 按行号排序
            items = sorted(items, key = lambda x: x.get_row())

            # 数据处理
            abs_path = f"{output_path}/双语对照/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with zipfile.ZipFile(self.insert_source_target(abs_path), "w") as zip_writer:
                with zipfile.ZipFile(f"{output_path}/cache/temp/{rel_path}", "r") as zip_reader:
                    for path in zip_reader.namelist():
                        if path.lower().endswith(".css"):
                            process_css(zip_reader, path)
                        elif path.lower().endswith(".opf"):
                            process_opf(zip_reader, path)
                        elif path.lower().endswith(".ncx"):
                            process_ncx(zip_reader, path, items)
                        elif path.lower().endswith((".html", ".xhtml")):
                            process_html(zip_reader, path, items, True)
                        else:
                            zip_writer.writestr(path, zip_reader.read(path))

    # # game/script8.rpy:16878
    # translate chinese arabialogoff_e5798d9a:
    #
    #     # "lo" "And you...?{w=2.3}{nw}" with dissolve
    #     # "lo" "" with dissolve
    #
    # # game/script/1-home/1-Perso_Home/elice.rpy:281
    # translate schinese elice_ask_home_f01e3240_5:
    #
    #     # e ".{w=0.5}.{w=0.5}.{w=0.5}{nw}"
    #     e ""
    #
    # # game/script8.rpy:33
    # translate chinese update08_a626b58f:
    #
    #     # "*Snorts* Fucking hell, I hate with this dumpster of a place." with dis06
    #     "" with dis06
    #
    # translate chinese strings:
    #
    #     # game/script8.rpy:307
    #     old "Accompany her to the inn"
    #     new ""
    #
    #     # game/script8.rpy:2173
    #     old "{sc=3}{size=44}Jump off the ship.{/sc}"
    #     new ""
    #
    # # game/routes/endings/laura/normal/Harry/l_normal_11_h.rpy:3
    # translate schinese l_normal_11_h_f9190bc9:
    #
    #     # nvl clear
    #     # n "After a wonderful night, the next day, to our displeasure, we were faced with the continuation of the commotion that I had accidentally engendered the morning prior."
    #     n ""

    # RENPY
    def read_from_path_renpy(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                lines = [line.removesuffix("\n") for line in reader.readlines()]

            for line in lines:
                results: list[str] = FileManager.RE_RENPY.findall(line)
                is_content_line = line.startswith("    # ") or line.startswith("    old ")

                # 不是内容行但找到匹配项目时，则直接跳过这一行
                if is_content_line == False and len(results) > 0:
                    continue
                elif is_content_line == True and len(results) == 1:
                    content = results[0].replace("\\n", "\n").replace("\\\"", "\"")
                elif is_content_line == True and len(results) >= 2:
                    content = results[1].replace("\\n", "\n").replace("\\\"", "\"")
                else:
                    content = ""

                # 添加数据
                items.append(
                    CacheItem({
                        "src": content,
                        "dst": content,
                        "extra_field": line,
                        "row": len(items),
                        "file_type": CacheItem.FileType.RENPY,
                        "file_path": rel_path,
                        "text_type": CacheItem.TextType.RENPY,
                    })
                )

        return items

    # RENPY
    def write_to_path_renpy(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:

        def repl(m: re.Match, i: list[int], t: int, dst: str) -> str:
            if i[0] == t:
                i[0] = i[0] + 1
                return f"\"{dst}\""
            else:
                i[0] = i[0] + 1
                return m.group(0)

        def process(text: str) -> str:
            return text.replace("\n", "\\n").replace("\\\"", "\"").replace("\"", "\\\"")

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.RENPY
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                line = item.get_extra_field()
                results: list[str] = FileManager.RE_RENPY.findall(line)

                i = [0]
                if line.startswith("    # "):
                    result.append(line)
                    if len(results) == 1:
                        line = FileManager.RE_RENPY.sub(lambda m: repl(m, i, 0, process(item.get_dst())), line)
                        result.append(f"    {line.removeprefix("    # ")}")
                    elif len(results) >= 2:
                        line = FileManager.RE_RENPY.sub(lambda m: repl(m, i, 1, process(item.get_dst())), line)
                        result.append(f"    {line.removeprefix("    # ")}")
                elif line.startswith("    old "):
                    result.append(line)
                    if len(results) == 1:
                        line = FileManager.RE_RENPY.sub(lambda m: repl(m, i, 0, process(item.get_dst())), line)
                        result.append(f"    new {line.removeprefix("    old ")}")
                    elif len(results) >= 2:
                        line = FileManager.RE_RENPY.sub(lambda m: repl(m, i, 1, process(item.get_dst())), line)
                        result.append(f"    new {line.removeprefix("    old ")}")
                else:
                    result.append(line)

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join(result))

    # KV JSON
    def read_from_path_kvjson(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        # {
        #     "「あ・・」": "「あ・・」",
        #     "「ごめん、ここ使う？」": "「ごめん、ここ使う？」",
        #     "「じゃあ・・私は帰るね」": "「じゃあ・・私は帰るね」",
        # }

        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                json_data: dict[str, str] = json.load(reader)

                # 格式校验
                if not isinstance(json_data, dict):
                    continue

                # 读取数据
                for k, v in json_data.items():
                    if isinstance(k, str) and isinstance(v, str):
                        # 根据是否已存在翻译数据来添加
                        if k.strip() != v.strip() or v.strip() == "":
                            items.append(
                                CacheItem({
                                    "src": k,
                                    "dst": v,
                                    "row": len(items),
                                    "file_type": CacheItem.FileType.KVJSON,
                                    "file_path": rel_path,
                                    "status": Base.TranslationStatus.EXCLUDED,
                                })
                            )
                        else:
                            items.append(
                                CacheItem({
                                    "src": k,
                                    "dst": v,
                                    "row": len(items),
                                    "file_type": CacheItem.FileType.KVJSON,
                                    "file_path": rel_path,
                                    "status": Base.TranslationStatus.UNTRANSLATED,
                                })
                            )

        return items

    # KV JSON
    def write_to_path_kvjson(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # {
        #     "「あ・・」": "「あ・・」",
        #     "「ごめん、ここ使う？」": "「ごめん、ここ使う？」",
        #     "「じゃあ・・私は帰るね」": "「じゃあ・・私は帰るね」",
        # }

        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.KVJSON
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write(
                    json.dumps(
                        {
                            item.get_src(): item.get_dst() for item in items
                        },
                        indent = 4,
                        ensure_ascii = False,
                    )
                )

    # Message JSON
    def read_from_path_messagejson(self, input_path: str, output_path: str, abs_paths: list[str]) -> list[CacheItem]:
        # [
        #     {
        #         "name", "しますか",
        #         "message": "<fgName:pipo-fog004><fgLoopX:1><fgLoopY:1><fgSx:-2><fgSy:0.5>"
        #     },
        #     {
        #         "message": "エンディングを変更しますか？"
        #     },
        #     {
        #         "message": "はい"
        #     },
        # ]

        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, input_path)

            # 数据处理
            with open(abs_path, "r", encoding = "utf-8-sig") as reader:
                json_data: list[dict] = json.load(reader)

                # 格式校验
                if not isinstance(json_data, list):
                    continue

                for v in json_data:
                    if isinstance(v, dict) and "message" in v:
                        items.append(
                            CacheItem({
                                "src": v.get("message", ""),
                                "dst": v.get("message", ""),
                                "extra_field": v.get("name", None),
                                "row": len(items),
                                "file_type": CacheItem.FileType.MESSAGEJSON,
                                "file_path": rel_path,
                            })
                        )

        return items

    # Message JSON
    def write_to_path_messagejson(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # [
        #     {
        #         "message": "<fgName:pipo-fog004><fgLoopX:1><fgLoopY:1><fgSx:-2><fgSy:0.5>"
        #     },
        #     {
        #         "message": "エンディングを変更しますか？"
        #     },
        #     {
        #         "message": "はい"
        #     },
        # ]

        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.MESSAGEJSON
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                if item.get_extra_field() is None:
                    result.append({
                        "message": item.get_dst(),
                    })
                else:
                    result.append({
                        "name": item.get_extra_field(),
                        "message": item.get_dst(),
                    })

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write(json.dumps(result, indent = 4, ensure_ascii = False))